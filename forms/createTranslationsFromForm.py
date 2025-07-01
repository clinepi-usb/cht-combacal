import openpyxl
import os.path


#get all the xlsx files in the forms directory /home/kevin/cht-core/config/cht-combacal-main-trial-config-code/forms/app
fileNames = []
my_path = os.path.abspath(os.path.dirname(__file__))
path = os.path.join(my_path, "../forms/app")
for file in os.listdir(path):
    if file.endswith(".xlsx"):
        fileNames.append(file)

# print all the file names
print("The following files were found:")
for fileName in fileNames:
    print(fileName)


# get the last row of the file with the most rows
for fileName in fileNames:
    path = os.path.join(my_path, "../forms/app/"+fileName)
    targetFile = openpyxl.load_workbook(path)
    target = targetFile.get_sheet_by_name('survey')
    if 'maxRow' in locals():
        if target.max_row > maxRow:
            maxRow = target.max_row
    else:
        maxRow = target.max_row

print("The file with the most rows has "+str(maxRow)+" rows")

# save the values in an array
names = []
# save the values in an array
labels = []

for fileName in fileNames:
    print("filename")
    print(fileName)
    path = os.path.join(my_path, "../forms/app/"+fileName)
    
    targetFile = openpyxl.load_workbook(path)
    target = targetFile.get_sheet_by_name('survey')
    # ignore the first row
    rowstart = 2
    rowend = maxRow

    # Iterate over the second column of the sheet
    # Create a stack to keep track of the groups
    # Each time in row i column 1 is a "begin group" we add value row i column 2 to the stack
    # Each time in row i column 1 is an "end group" we pop the last value from the stack
    stack = []
    for i in range(rowstart, rowend+1):
        # If in column 1 is a "begin group," we have to add the value to the stack
        # and also the value in column 2 to names and the value in column 3 to labels
        name = "report." + fileName.split(".")[0] + "."
        label = ""
        # don't translate data group
        if target.cell(row=i, column=1).value == "begin group" and target.cell(row=i, column=2).value == "data":
            break
        # skip columns where name and label is nonType go to next row
        if (target.cell(row=i, column=1).value is None and target.cell(row=i, column=2).value is None):
            continue
        if target.cell(row=i, column=1).value == "begin group":
            stack.append(target.cell(row=i, column=2).value)
            # if begin group "inputs", skip
            if stack[0] == "inputs":
                continue

            if target.cell(row=i, column=1).value == "begin group" and target.cell(row=i, column=2).value == "inputs":
                continue
            if len(stack) == 1 and target.cell(row=i, column=2).value not in names and target.cell(row=i, column=2).value is not None:
                name = name + target.cell(row=i, column=2).value
                label = target.cell(row=i, column=3).value
                if target.cell(row=i, column=3).value is None:
                    label = "No label"
                if name == "report.vhw_clin_ev.buk.today":
                    print("here1")
                if name not in names:
                    names.append(name)
                    labels.append(label)
            if len(stack) > 1 and target.cell(row=i, column=2).value not in names and target.cell(row=i, column=2).value is not None:
                for j in range(0, len(stack)-1):
                    name += stack[j]+"."
                name += target.cell(row=i, column=2).value
                label = target.cell(row=i, column=3).value
                if target.cell(row=i, column=3).value is None:
                    label = "No label"
                if name == "report.vhw_clin_ev.buk.today":
                    print("here2")
                if name not in names:
                    names.append(name)
                    labels.append(label)
        elif target.cell(row=i, column=1).value == "end group":
            stack.pop()
        elif target.cell(row=i, column=2).value is not None:  # Check if the value is not None
            # if in group "inputs" skip
            if len(stack) >= 1 and stack[0] == "inputs":
                continue
            if len(stack) == 0:
                name = name + target.cell(row=i, column=2).value
                label = target.cell(row=i, column=3).value
            else:
                for j in range(0, len(stack)):
                    name += stack[j]+"."
                name += target.cell(row=i, column=2).value
                label = target.cell(row=i, column=3).value
            if target.cell(row=i, column=3).value is None:
                label = "No label"
            if name == "report.vhw_clin_ev.buk.today":
                    print("here3")
            if name not in names:
                names.append(name)
                labels.append(label)
            
    # print file name
    print("File: "+fileName)

    # go though names and labels
    #if report.supvis_clin_ev.t_aht_txs = NO_LABEL take the subword after the first t_ and add it to "Preloaded "
    for i in range(0, len(names)-1):
        # check if there is a t_ in the name
        if "t_" in names[i]:
            # find the index of the first .t_
            index = names[i].index("t_")
            # add the subword after the first .t_ to "Preloaded "
            labels[i] = "Preloaded "+names[i][index+2:]

        if ".patient_uuid" in names[i]:
            labels[i] = "Patient UUID"
        if ".patient_id" in names[i]:
            labels[i] = "Patient ID"
        if ".patient_name" in names[i]:
            labels[i] = "Patient Name"
        if ".age" in names[i]:
            labels[i] = "Age"
        if ".sex" in names[i]:
            labels[i] = "Sex"
        if ".twic_arm" in names[i]:
            labels[i] = "TWIC ARM"

        # if there is one or more " " in the labels and nothing else, use the subword after the last . of the name
        if " " in labels[i] and labels[i].count(" ") == len(labels[i]):
            # find the index of the last .
            index = names[i].rindex(".")
            # add the subword after the last . to the label
            labels[i] = names[i][index+1:]

        if labels[i] == " ":
            # find the index of the last .
            index = names[i].rindex(".")
            # add the subword after the last . to the label
            labels[i] = names[i][index+1:]

        # if there is a NO_LABEL in the labels, use the subword after the last . of the name
        if labels[i] == "NO_LABEL" or labels[i] == "No label":
            # find the index of the last .
            index = names[i].rindex(".")
            # add the subword after the last . to the label
            labels[i] = names[i][index+1:]

        if names[i] == "report.dm_cont.date_clin_ev_rep":
            print(names[i])
            print("label:"+labels[i]+"end")
            print("length of labels: "+str(len(labels[i])))

    

    # add the values and the labels to a text file
    # open the file
    file = open("translations.txt", "w")
    for i in range(0, len(names)-1):
        # if there is an error in names[i]+" = "+labels[i] print names[i] and labels[i]
        if names[i] is None or labels[i] is None:
            print(names[i])
            print(labels[i])

        file.write(names[i]+" = "+labels[i]+"\n")
