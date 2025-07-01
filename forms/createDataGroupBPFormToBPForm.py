import openpyxl
import os.path


#choose what file to edit
fileName = input("Write the exact name of the form you would like to create a data group in\n")

my_path = os.path.abspath(os.path.dirname(__file__))
path = os.path.join(my_path, "../forms/app/"+fileName)
  
targetFile = openpyxl.load_workbook(path)
target = targetFile.get_sheet_by_name('survey')

#Deside on what row you want to create a group data.
#If there is already a group data, give the row
print("Is there already a group data?\n")
dataGroup = int(input("Yes: 1\nNo: 0\n"))

if (dataGroup == 1):
  rowNumber = int(input("In what row is the group data?\n"))
elif (dataGroup == 0):
  rowNumber = int(input("On what row should I create a data group?\n"))

  # make sure that there is no entry in and below the row rowNumber
  # if there is, ask again
  for i in range(rowNumber, target.max_row+1):
    if (target.cell(row=i, column=1).value != None or target.cell(row=i, column=2).value != None):
      print("There is already an entry in row "+str(rowNumber)+" or below")
      rowNumber = int(input("On what row should I create a data group?\n"))
      break

# if another number is given, ask again
else:
  print("Please enter either 1 or 0")
  dataGroup = int(input("Yes: 1\nNo: 0\n"))

#create a string of all the preloadingvariables
preloadingVariables = []

#iterate over the second column of the sheet
# insert the values after the value preloads and before the value preloadsend into the list
for i in range(1, target.max_row+1):
  if (target.cell(row=i, column=2).value == "preloads"):
    for j in range(i+1, target.max_row+1):
      # print the value of the cell
      # stop when the value is preloadsend
      if (target.cell(row=j, column=2).value == "preloads"):
        break
      else:
        # if the value starts with t_ it and is not a none value, add it to the list
        if (target.cell(row=j, column=2).value != None and target.cell(row=j, column=2).value.startswith("t_")):
        # take away the t_ from the value
          preloadingVariables.append(target.cell(row=j, column=2).value[2:])

# on column rowNumber create a group data
# insert begin group in cell 1, data in cell 2
target.cell(row=rowNumber, column=1).value = "begin group"
target.cell(row=rowNumber, column=2).value = "data"


# for every value in the list, check if it is in the column 2
# if it is, insert the value in column 2 in the next row
for i in range(0, len(preloadingVariables)):
  for j in range(1, target.max_row+1):
    if (target.cell(row=j, column=2).value == preloadingVariables[i]):
      rowNumber += 1
      # insert the value hidden in the first column after the row rowNumber
      target.cell(row=rowNumber, column=1).value = "hidden"
      # insert the value in the second column after the row rowNumber
      target.cell(row=rowNumber, column=2).value = "_"+preloadingVariables[i]
      # insert the value NO_LABEL in the third column after the row rowNumber
      target.cell(row=rowNumber, column=3).value = "NO_LABEL"

      # insert the value ${preloadingVariables[i]} in the tenth column after the row rowNumber
      target.cell(row=rowNumber, column=10).value = "${"+preloadingVariables[i]+"}"

# insert end group in cell 1, data in cell 2
rowNumber += 1
target.cell(row=rowNumber, column=1).value = "end group"
target.cell(row=rowNumber, column=2).value = "data"

# save the file
targetFile.save(path)



  
    