import openpyxl
import os.path


# Ask for the name of the excel file
fileName = input("What is the name of the excel file?\n")

# load the excel file
my_path = os.path.abspath(os.path.dirname(__file__))
path = os.path.join(my_path, "../forms/"+fileName)
  
targetFile = openpyxl.load_workbook(path)
#get first sheet
target = targetFile.get_sheet_by_name('village_list(3)')

# get the number of rows
maxRow = target.max_row + 1

# ask if "cont" or "int" villages should be extracted
print("Do you want to extract the cont or int villages?\n")
villageType = input("cont: 1\nint: 2\n")


# go through all the rows and copy the uuid in the format "0238d464-898c-44c3-87f4-1697c17bfe62" and "int" or "cont" in the format "int" or "cont"
villages = []
for i in range(2, maxRow):
    if (target.cell(row=i, column=1).value != None):
        # take out the value of the cell
        string = str(target.cell(row=i, column=1).value)
        # this is in string: 1,"0238d464-898c-44c3-87f4-1697c17bfe62","cont","bb","muelahc","khatleng",53533
        # split the string at the commas
        string = string.split(",")
        # this is in string: ['1', '"0238d464-898c-44c3-87f4-1697c17bfe62"', '"cont"', '"bb"', '"muelahc"', '"khatleng"', '53533']
        # take out the uuid and "int" or "cont"
        string = string[1:3]

        # delete the quotation marks
        string[0] = string[0].replace('"', '')
        string[1] = string[1].replace('"', '')
        
        # if the villageType is "cont" keep only the "cont" villages and if the villageType is "int" keep only the "int" villages
        if (villageType == "1" and string[1] == "cont"):
            villages.append(string)
        elif (villageType == "2" and string[1] == "int"):
            villages.append(string)


# for each village create this string and print it contact.contact.parent.parent._id === '04417203-80df-4b6d-948e-d7f6ab5a2501' ||
for i in range(0, len(villages)):
    print("contact.contact.parent.parent._id === '"+villages[i][0]+"' ||")