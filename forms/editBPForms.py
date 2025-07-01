import openpyxl
import os.path


#choose what file to edit
print('Enter the number of the file you would like to edit:\n')
fileNumber = input('{0} :1\n{1} :2\n{2} :3\n{3} :4\n{4} :5\n{5} :6\n{6} :7\n '.format('bp_confirm.xlsx', 'bp_followup_cont.xlsx', 'bp_followup_int.xlsx', 'bp_initiate.xlsx', 'bp_screening.xlsx', 'bp_updosing.xlsx', 'test_file.xlsx'))

if fileNumber == 7:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/test_file.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('Sheet1')

if fileNumber == 1:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_confirm.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('survey')

if fileNumber == 2:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_followup_cont.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('survey')

if fileNumber == 3:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_followup_int.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('survey')

if fileNumber == 4:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_initiate.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('survey')

if fileNumber == 5:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_screening.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('survey')

if fileNumber == 6:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_updosing.xlsx")
  
  targetFile = openpyxl.load_workbook(path)
  target = targetFile.get_sheet_by_name('survey')

# get sheet from templateFile
my_path = os.path.abspath(os.path.dirname(__file__))
path = os.path.join(my_path, "../forms/bpPreloadingTemplate.xlsx")
yfile = openpyxl.load_workbook(path)
template = yfile.get_sheet_by_name('Sheet1')




#Insert rows 1 to 28 from templateFile (until trial_arm) into target file if they are not yet available
templateRow = 0
for row in template.iter_rows(min_row = 1, max_col = 16, max_row = 28):
  templateRow += 1

  
  if template.cell(row = templateRow, column = 1).value != '' and template.cell(row = templateRow, column = 1).value != target.cell(row = templateRow, column = 1).value or template.cell(row = templateRow, column = 2).value != target.cell(row = templateRow, column = 2).value:
    target.insert_rows(templateRow)
    templateColumn = 1
    for cell in row:
      target.cell(row = templateRow, column = templateColumn, value = cell.value)
      templateColumn += 1


#Delete rows until secon trial arm
targetRow = 28
deleteRows = 0
for row in target.iter_rows(min_row = 29, max_col = 16, max_row = 38):
  targetRow +=1
  deleteRows +=1
  if target.cell(row = targetRow, column = 2).value == 'trial_arm':
    target.delete_rows(28, deleteRows)
    break


# find out how many preloading variables there is
targetSettings = targetFile.get_sheet_by_name('settings')
fileName = targetSettings.cell(row = 2, column = 1).value
preloadingRows = input('Enter how many rows are filled with preloading variables in file {0}:  '.format(fileName))

#find out in what row the trial_arm variable is
trialArmRow = 26
for row in target.iter_rows(min_row = 27, max_col = 16, max_row = 38):
  trialArmRow +=1
  if target.cell(row = trialArmRow, column = 2).value == 'trial_arm':
    break

print('trialArmRow', trialArmRow)

#Enter calculate formula for trial arm
target.cell(row = trialArmRow, column = 12, value = '../inputs/contact/parent/parent/cluster_trial_arm')

#Make all fields hidden in column tpye for the preloading variables
targetRow = trialArmRow - 1

for row in target.iter_rows(min_row = trialArmRow, max_col = 16, max_row = trialArmRow + preloadingRows -1):
  targetRow +=1

  #Make all fields hidden in column tpye for the preloading variables
  target.cell(row = targetRow, column = 1, value = 'hidden')

  #Make all fields but the one next to triel_arm empty in column calculation for the preloading variables
  if targetRow > trialArmRow:
    target.cell(row = targetRow, column = 12, value = '')

  #Make all fields NO_LABEL in column label::en for the preloading variables
  target.cell(row = targetRow, column = 3, value = 'NO_LABEL')

  #Make all fields hidden in columns hint::en and hint::se for the preloading variables
  target.cell(row = targetRow, column = 14, value = 'hidden')
  target.cell(row = targetRow, column = 15, value = 'hidden')
  
#save eddited file
if fileNumber == 7:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/test_file.xlsx")
  
  targetFile.save(path)


if fileNumber == 1:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_confirm.xlsx")
  
  targetFile.save(path)
  

if fileNumber == 2:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_followup_cont.xlsx")
  
  targetFile.save(path)
  
 

if fileNumber == 3:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_followup_int.xlsx")
  
  targetFile.save(path)
  
 

if fileNumber == 4:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_initiate.xlsx")
  
  targetFile.save(path)
  

if fileNumber == 5:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_screening.xlsx")
  
  targetFile.save(path)
  

if fileNumber == 6:
  my_path = os.path.abspath(os.path.dirname(__file__))
  path = os.path.join(my_path, "../forms/app/bp_updosing.xlsx")
  
  targetFile.save(path)
 
